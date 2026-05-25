from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List

from article_tracker.models.article import Article


def write_excel(articles: List[Article], out_dir: str, prefix: str = "papers") -> str | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        return None

    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = d / f"{prefix}_{ts}.xlsx"

    wb = Workbook()
    tiers = {}
    for a in articles:
        tier = a.screening_tier.value if a.screening_tier else "unclassified"
        tiers.setdefault(tier, []).append(a)

    for tier_name, group in tiers.items():
        ws = wb.create_sheet(title=tier_name[:31])
        headers = ["Title", "Authors", "Venue", "Published", "DOI", "Tier", "Abstract", "Summary", "Code Links"]
        ws.append(headers)
        for a in group:
            ws.append([
                a.title,
                ", ".join(a.authors[:5]),
                a.venue or "",
                a.published or "",
                a.doi or "",
                tier_name,
                (a.abstract or "")[:200],
                a.digest_en or a.digest_zh or "",
                " | ".join(a.code_links[:3]),
            ])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(str(path))
    return str(path)
